from odoo import models, fields, api
from odoo.exceptions import ValidationError
import base64
from io import BytesIO
import openpyxl
from openpyxl import Workbook 
import logging

_logger = logging.getLogger(__name__)

class ImportStockQuantWizard(models.TransientModel):
    _name = 'import.stock.quant.wizard'
    _description = 'Wizard to adjust stock.quant using inventory_quantity + action_apply_inventory'

    file = fields.Binary(
        string="Excel File",
        help="Upload the XLSX file to assign inventory_quantity and apply with action_apply_inventory."
    )
    filename = fields.Char("File Name")

    # Nuevo campo para que el usuario elija la ubicación desde el wizard
    location_id_2 = fields.Many2one(
        'stock.location',
        string='Location',
        help='Select the location where this inventory adjustment will be applied.'
    )
        # Campo para GUARDAR el Excel a exportar
    export_file = fields.Binary("Product File", readonly=True)
    export_filename = fields.Char("Name to download", readonly=True)

    def action_export_products(self):
        """
        Create an Excel file with the columns: DEFAULT_CODE | NAME | QUANTITY
        and save in 'export_file' like base64.
        """
        _logger.info("===== [EXPORT PRODUCTS] Generating Excel of Products... =====")

        # 1) Obtener la lista de productos que quieras exportar
        #    (ejemplo: todos los productos que tengan default_code)
        products = self.env['product.product'].search([('detailed_type', '=', 'product')])

        # 2) Crear el archivo Excel con openpyxl
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"

        # Escribir cabeceras
        ws["A1"] = "DEFAULT_CODE"
        ws["B1"] = "NAME"
        ws["C1"] = "QUANTITY"

        # Escribir filas
        row = 2
        for product in products:
            ws.cell(row=row, column=1).value = product.default_code or ""
            ws.cell(row=row, column=2).value = product.name or ""
            # Cantidad disponible (puedes usar qty_available, virtual_available, etc.)
            ws.cell(row=row, column=3).value = product.qty_available  
            row += 1

        # 3) Guardar el Excel en memoria
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)

        # 4) Convertir a Base64 y asignar a 'export_file'
        excel_data = base64.b64encode(fp.read())
        fp.close()

        # 5) Asignar al wizard y dar un nombre de ejemplo
        self.export_file = excel_data
        self.export_filename = "Products.xlsx"

        _logger.info("===== [EXPORT PRODUCTS] Excel generated =====")

        # 6) Retornar la vista actual para que el usuario vea el nuevo archivo
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'import.stock.quant.wizard',
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }

    def action_import_stock_quant(self):
        """
        default_code (product)

        inventory_qty (counted quantity)
        Use self.location_id_2 as the destination location (instead of reading it from the Excel file).
        """
        _logger.info("===== [IMPORT STOCK QUANT] Adjustment with inventory_quantity + action_apply_inventory... =====")

        # Validar que sea XLSX
        if not self.filename or not self.filename.lower().endswith('.xlsx'):
            raise ValidationError("The file must be .xlsx")

        _logger.info(f"Trying to read the file: {self.filename}")
        try:
            data = base64.b64decode(self.file)
            file_input = BytesIO(data)
            workbook = openpyxl.load_workbook(file_input, data_only=True)
        except Exception as e:
            raise ValidationError(f"Can't be opened the XLSX file. Error: {e}")

        sheet = workbook.active
        _logger.info(f"XLSX file read successfully. Active sheet selected: {sheet.title}")

        filas_procesadas = 0
        for row_idx in range(2, sheet.max_row + 1):
            # Si la fila está en blanco, salimos
            if all(sheet.cell(row=row_idx, column=col).value in [None, ""]
                   for col in range(1, sheet.max_column + 1)):
                break

            # Columna 1 => default_code
            default_code = sheet.cell(row=row_idx, column=1).value  
            # Columna 3 => inventory_qty
            new_inventory_qty = sheet.cell(row=row_idx, column=3).value  

            # Validación básica
            if not default_code:
                _logger.warning(f"[ROW {row_idx}] Falta default_code (producto). Se omite.")
                continue

            # Convertir la cantidad a float
            if isinstance(new_inventory_qty, str):
                try:
                    new_inventory_qty = float(new_inventory_qty.replace(',', '.'))
                except ValueError:
                    _logger.warning(f"[ROW {row_idx}] can't converted '{new_inventory_qty}' to float. Omit.")
                    continue

            # Buscar el producto por default_code
            product = self.env['product.product'].search([('default_code', '=', str(default_code))], limit=1)
            if not product:
                _logger.warning(f"[ROW {row_idx}] Product '{default_code}' not found. Omite.")
                continue

            # Usar la ubicación seleccionada en el wizard
            location = self.location_id_2

            # Buscar o crear quant
            quant = self.env['stock.quant'].search([
                ('product_id', '=', product.id),
                ('location_id', '=', location.id),
            ], limit=1)
            if not quant:
                quant = self.env['stock.quant'].create({
                    'product_id': product.id,
                    'location_id': location.id,
                    'quantity': 0.0,
                })
                _logger.info(f"[ROW {row_idx}] Created quant({quant.id}) for {default_code} in loc={location.id}")

            # Asignar inventory_quantity
            quant.inventory_quantity = new_inventory_qty
            quant.inventory_quantity_set = True

            # Aplicar el inventario => genera moves/lines
            _logger.info(f"[ROW {row_idx}] Calling quant({quant.id}).action_apply_inventory() con inv_qty={new_inventory_qty}...")
            quant.action_apply_inventory()

            filas_procesadas += 1
            _logger.info(f"[ROW {row_idx}] Adjustment aplicated in ({quant.id}).")

        _logger.info(f"===== [IMPORT STOCK QUANT] Finalized. Processed ROWs: {filas_procesadas} =====")
        return {'type': 'ir.actions.client', 'tag': 'reload'}
